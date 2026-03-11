import logging
from openpyxl import load_workbook
from typing import Any
from configparser import ConfigParser
from pathlib import Path

from loaders.parameter_preprocessing import is_missing

PlanetStarDict = dict[str, Any]

def load_matching_excel_row_from_excel(excel_path, target_name_user_input):
    try:

        workbook = load_workbook(excel_path, data_only=True)
        logging.info("Opening Excel file: %s", excel_path)

        # We currently assume that the **active worksheet** is the one containing the parameters we care about.
        worksheet = workbook.active

        #normalize headers so we can match the name
        column_headers = []
        for cell in worksheet[1]:
            name = _strip_hash_from_excel_column(cell.value)
            if name is None:
                break
            column_headers.append(name)
        logging.info("Excel column headers: %s", column_headers)

        if "pl_name" not in column_headers:
            raise ValueError("Excel file has no 'pl_name' column")

        # Find matching row by pl_name
        pl_name_index = column_headers.index("pl_name")
        target_name_normalized = str(target_name_user_input).casefold()
        logging.info("Searching for target: '%s'", target_name_user_input)


        matching_row_dict = None
        checked_names = []

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            # Guard: rows shorter than headers: i also want to get null values for empty cells
            row = list(row) + [None] * (len(column_headers) - len(row))

            pl_name_value = row[pl_name_index]
            if pl_name_value is None:
                break

            pl_name_normalized = str(pl_name_value).casefold().strip()
            checked_names.append(pl_name_normalized)

            # First match wins. :(
            # TODO: harden later when planetary transits are a thing.
            # Requirement: strip, drop last character, strip again
            pl_base = pl_name_normalized[:-1].strip() if pl_name_normalized else ""

            if pl_base == target_name_normalized:
                matching_row_dict = {
                    header: value
                    for header, value in zip(column_headers, row)
                    if header is not None
                }
                logging.info("Found matching row for target '%s':", target_name_user_input)
                break

        if matching_row_dict is None:
            logging.error("All pl_name values checked: %s", checked_names)
            raise ValueError(
                f"No target found for '{target_name_user_input}' (searched until row with empty pl_name)"
            )

        return matching_row_dict, target_name_user_input
    except Exception:
        logging.exception("Failed to read matching Excel row for target_name_user_input=%r from excel_path=%s", target_name_user_input, excel_path)
        raise

def load_excel_cfg(mapping_path: Path):
    """
    Load Excel-to-canonical mapping configuration.

    Returns a dict with:
      - planet: {canonical_key: excel_header}
      - star: {canonical_key: excel_header}
      - required_planetary_parameters: [canonical_key, ...]
      - required_stellar_parameters: [canonical_key, ...]
    """
    try:

        cfg = ConfigParser()
        # preserve case of canonical keys
        cfg.optionxform = str

        mapping_path = Path(mapping_path)
        if not mapping_path.exists():
            raise FileNotFoundError(f"Excel mapping file not found: {mapping_path}")

        cfg.read(mapping_path)
        logging.info("Loaded Excel mapping from %s", mapping_path)
        logging.info("Config sections found: %s", cfg.sections())

        # Parse required keys; missing section means "no required keys".
        def parse_required(section: str):
            if not cfg.has_section(section):
                return []
            raw = cfg.get(section, "keys", fallback="")
            return [k.strip() for k in raw.split(",") if k.strip()]


        mapping = {
            "planet": dict(cfg.items("planets")) if cfg.has_section("planets") else {},
            "star": dict(cfg.items("stars")) if cfg.has_section("stars") else {},
            "required_planetary_parameters": parse_required("required_planetary_parameters"),
            "required_stellar_parameters": parse_required("required_stellar_parameters"),
        }

        logging.info("Excel mapping: %d planet keys, %d star keys, %d required planet, %d required star", len(mapping["planet"]), len(mapping["star"]), len(mapping["required_planetary_parameters"]), len(mapping["required_stellar_parameters"]))
        logging.info("Planet mapping keys: %s", mapping["planet"].keys())
        logging.info("Star mapping keys: %s", mapping["star"].keys())

        return mapping

    except Exception:
        logging.exception("Failed to load Excel mapping cfg from %s", mapping_path)
        raise


def map_to_planet_or_star_dictionary(planet_star_dictionary: PlanetStarDict, mapping: dict, target_name: str) -> tuple[dict[str, Any], dict[str, Any]]:
    """
    Translate one Excel row dict (Excel headers -> values) into two dicts:
      planet_params: canonical keys -> values
      star_params: canonical keys -> values

    Uses mapping loaded by load_excel_cfg().
    Validates required keys from [required_planetary_parameters] and [required_stellar_parameters].
    """
    # normalize if excel columns have leading or trailing spaces or upper lower case chars
    def norm(s: str) -> str:
        return str(s).strip().casefold()

    # Build reverse lookup: excel_header -> canonical_key
    reverse_planets = {norm(excel_header): canonical for canonical, excel_header in mapping["planet"].items()}
    reverse_star = {norm(excel_header): canonical for canonical, excel_header in mapping["star"].items()}

    planet_params: PlanetStarDict = {}
    star_params: PlanetStarDict = {}

    unknown_headers: list[str] = []
    collisions: list[str] = []

    for excel_header, value in planet_star_dictionary.items():
        h = norm(excel_header)

        if h in reverse_planets:
            canonical = reverse_planets[h]
            if canonical in planet_params:
                collisions.append(f"planet:{canonical}")
            planet_params[canonical] = value
            continue

        if h in reverse_star:
            canonical = reverse_star[h]
            if canonical in star_params:
                collisions.append(f"star:{canonical}")
            star_params[canonical] = value
            continue

        unknown_headers.append(str(excel_header))

    if collisions:
        raise ValueError(f"Multiple Excel columns mapped to the same canonical key: {sorted(set(collisions))}")

    # i don't care about not mapped columns. I just write it to the log and carry on.
    if unknown_headers:
        logging.warning("Ignoring unmapped Excel columns:\n%s", "\n".join(f"  {h!r}" for h in unknown_headers))
    
    # no star's name in excel, so we insert the matched star name from the row 
    star_params["name"] = target_name

    missing_planet = [k for k in mapping["required_planetary_parameters"] if k not in planet_params or is_missing(planet_params[k])]
    missing_star = [k for k in mapping["required_stellar_parameters"] if k not in star_params or is_missing(star_params[k])]
    logging.info("Missing required planet properties after Excel import: %s", missing_planet)
    logging.info("Missing required star properties after Excel import: %s", missing_star)

    logging.info("Processed planetary parameters after Excel import: %s", planet_params)
    logging.info("Processed stellar parameters after Excel import: %s", star_params)

    return planet_params, star_params

def _strip_hash_from_excel_column(name):
    """Normalize a column header: strip whitespace and remove a leading '#'."""
    if name is None:
        return None
    name = str(name).strip()
    if name.startswith("#"):
        name = name[1:]
    return name
