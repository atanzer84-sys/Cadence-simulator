import logging
from openpyxl import load_workbook
from typing import Any
from configparser import ConfigParser
from pathlib import Path
PlanetStarDict = dict[str, Any]

def load_excel_parameters(excel_path, target_name_user_input):
    workbook = load_workbook(excel_path, data_only=True)

    # We currently assume that the **active worksheet** is the one containing the parameters we care about.
    worksheet = workbook.active

    #normalize headers so we can match the name
    column_headers = [_normalize_name(cell.value) for cell in worksheet[1]] 
    logging.info("Excel column headers: %s", column_headers)

    if "pl_name" not in column_headers:
        raise ValueError("Excel file has no 'pl_name' column")

    # Find matching row by pl_name
    pl_name_index = column_headers.index("pl_name")
    target_name = str(target_name_user_input).strip()
    target_name_normalized = str(target_name).casefold()
    logging.info("Searching for target: '%s'", target_name)

    matching_row_dict = None
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        # Guard: rows shorter than headers: i also want to get null values for empty cells
        row = list(row) + [None] * (len(column_headers) - len(row))

        pl_name_value = row[pl_name_index]
        if pl_name_value is None:
            break

        pl_name_normalized = str(pl_name_value).casefold().strip()

        # First match wins. :(
        # TODO: harden later when planetary transits are a thing.
        if pl_name_normalized.startswith(target_name_normalized):
            matching_row_dict = {
                header: value
                for header, value in zip(column_headers, row)
                if header is not None
            }
            logging.info("Found matching row for target '%s':", target_name_user_input)
            break

    if matching_row_dict is None:
        raise ValueError(
            f"No target found for '{target_name_user_input}' (searched until row with empty pl_name)"
        )
    return matching_row_dict, target_name

def load_excel_mapping(mapping_path: Path):
    """
    Load Excel-to-canonical mapping configuration.

    Returns a dict with:
      - planet: {canonical_key: excel_header}
      - star: {canonical_key: excel_header}
      - required_planet_parameters: [canonical_key, ...]
      - required_star_parameters: [canonical_key, ...]
    """
    cfg = ConfigParser()
    # preserve case of canonical keys
    cfg.optionxform = str

    mapping_path = Path(mapping_path)
    if not mapping_path.exists():
        raise FileNotFoundError(f"Excel mapping file not found: {mapping_path}")

    cfg.read(mapping_path)
    logging.info("Loaded Excel mapping from %s", mapping_path)
    logging.info("Config sections found: %s", cfg.sections())

    # Parse required keys; missing section means "no required keys".
    def parse_required(section: str):
        if not cfg.has_section(section):
            return []
        raw = cfg.get(section, "keys", fallback="")
        return [k.strip() for k in raw.split(",") if k.strip()]


    mapping = {
        "planet": dict(cfg.items("planets")) if cfg.has_section("planets") else {},
        "star": dict(cfg.items("stars")) if cfg.has_section("stars") else {},
        "required_planet_parameters": parse_required("required_planet_parameters"),
        "required_star_parameters": parse_required("required_star_parameters"),
    }

    logging.info(
        "Excel mapping: %d planet keys, %d star keys, %d required planet, %d required star",
        len(mapping["planet"]),
        len(mapping["star"]),
        len(mapping["required_planet_parameters"]),
        len(mapping["required_star_parameters"]),
    )
    logging.info("Planet mapping keys: %s", mapping["planet"].keys())
    logging.info("Star mapping keys: %s", mapping["star"].keys())

    return mapping

def map_excel_row(planet_star_dictionary: PlanetStarDict, mapping: dict, target_name: str) -> tuple[dict[str, Any], dict[str, Any]]:
    """
    Translate one Excel row dict (Excel headers -> values) into two dicts:
      planet_params: canonical keys -> values
      star_params: canonical keys -> values

    Uses mapping loaded by load_excel_mapping().
    Validates required keys from [required_planet_parameters] and [required_star_parameters].
    """
    # normalize if excel columns have leading or trailing spaces or upper lower case chars
    def norm(s: str) -> str:
        return str(s).strip().casefold()

    # mandatory checks for required params
    def is_missing(v: Any) -> bool:
        if v is None:
            return True
        if isinstance(v, str) and v.strip() == "":
            return True
        return False

    # Build reverse lookup: excel_header -> canonical_key
    reverse_planets = {norm(excel_header): canonical for canonical, excel_header in mapping["planet"].items()}
    reverse_star = {norm(excel_header): canonical for canonical, excel_header in mapping["star"].items()}

    planet_params: PlanetStarDict = {}
    star_params: PlanetStarDict = {}

    unknown_headers: list[str] = []
    collisions: list[str] = []

    for excel_header, value in planet_star_dictionary.items():
        h = norm(excel_header)

        if h in reverse_planets:
            canonical = reverse_planets[h]
            if canonical in planet_params:
                collisions.append(f"planet:{canonical}")
            planet_params[canonical] = value
            continue

        if h in reverse_star:
            canonical = reverse_star[h]
            if canonical in star_params:
                collisions.append(f"star:{canonical}")
            star_params[canonical] = value
            continue

        unknown_headers.append(str(excel_header))

    if collisions:
        raise ValueError(f"Multiple Excel columns mapped to the same canonical key: {sorted(set(collisions))}")

    # i don't care about not mapped columns. I just write it to the log and carry on.
    if unknown_headers:
        logging.warning(
            "Ignoring unmapped Excel columns:\n%s",
            "\n".join(f"  {h!r}" for h in unknown_headers),
        )
    
    # no star's name in excel, so we insert the matched star name from the row 
    star_params["name"] = target_name

    missing_planet = [k for k in mapping["required_planet_parameters"] if k not in planet_params or is_missing(planet_params[k])]
    missing_star = [k for k in mapping["required_star_parameters"] if k not in star_params or is_missing(star_params[k])]
    logging.info("Missing required planet params: %s", missing_planet)
    logging.info("Missing required star params: %s", missing_star)

    # if missing_planet:
    #     raise ValueError(f"Missing required planet parameters: {missing_planet}")

    # if missing_star:
    #     raise ValueError(f"Missing required star parameters: {missing_star}")

    logging.info("Processed planetary parameters:")
    for k, v in planet_params.items():
        logging.info("  %s = %r", k, v)

    logging.info("Processed stellar parameters:")
    for k, v in star_params.items():
        logging.info("  %s = %r", k, v)

    return planet_params, star_params

def _normalize_name(name):
    """Normalize a column header: strip whitespace and remove a leading '#'."""
    if name is None:
        return None
    name = str(name).strip()
    if name.startswith("#"):
        name = name[1:]
    return name