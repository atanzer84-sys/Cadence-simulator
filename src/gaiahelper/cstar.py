import re
import csv
from pathlib import Path
from loaders.run_waltzer_context import setup_output_directory
from openpyxl import Workbook
from openpyxl.utils import get_column_letter



photometry_pattern = re.compile(
    r"Aperture photometry \((?P<channel>[^)]+)\):.*?"
    r"Cc=(?P<Cc>[-+0-9.eE]+)\s+"
    r"Ca=(?P<Ca>[-+0-9.eE]+)\s+"
    r"Nc=(?P<Nc>\d+)\s+"
    r"Na=(?P<Na>\d+)\s+"
    r"C_background=(?P<C_background>[-+0-9.eE]+)\s+"
    r"C_star=(?P<C_star>[-+0-9.eE]+)\s+"
    r"C_star_noise=(?P<C_star_noise>[-+0-9.eE]+)"
)

target_pattern = re.compile(
    r"UserConfig\(target_name='(?P<target>[^']+)'"
)

exposure_pattern = re.compile(
    r"Channel (?P<channel>\w+) science frame calculation:.*?exposure_s=(?P<exposure>[0-9.]+)"
)

parameter_file_pattern = re.compile(
    r"User parameter file loaded: (?P<parameter_file>.+)"
)

def format_value_for_csv(v):
    if isinstance(v, float):
        return str(v).replace(".", ",")
    return v

def csv_to_excel(csv_path, xlsx_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "photometry_summary"

    with csv_path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f, delimiter=";")
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)

                if row_idx == 1:
                    cell.value = value
                    continue

                if isinstance(value, str):
                    stripped = value.strip()
                    try:
                        if "," in stripped:
                            number = float(stripped.replace(",", "."))
                            cell.value = number
                        else:
                            number = int(stripped)
                            cell.value = number
                    except ValueError:
                        cell.value = value
                else:
                    cell.value = value

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            cell_str = "" if cell.value is None else str(cell.value)
            if len(cell_str) > max_length:
                max_length = len(cell_str)
        ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(xlsx_path)



def main(existing_output_dir=None):

    if existing_output_dir is None:
        log_dir, _, _ = setup_output_directory()
    else:
        log_dir = Path(existing_output_dir)

    out_csv = log_dir / "photometry_summary.csv"
    out_xlsx = log_dir / "photometry_summary.xlsx"

    rows = []

    for log_file in sorted(log_dir.rglob("*.log")):

        target_name = None
        exposures = {}
        frame_counters = {}

        with log_file.open("r", encoding="utf-8", errors="ignore") as f:
            for line_number, line in enumerate(f, start=1):

                t = target_pattern.search(line)
                if t:
                    target_name = t.group("target")

                e = exposure_pattern.search(line)
                if e:
                    exposures[e.group("channel")] = float(e.group("exposure"))

                m = photometry_pattern.search(line)
                if not m:
                    continue

                channel = m.group("channel")

                key = (log_file.parent.name, channel)

                if key not in frame_counters:
                    frame_counters[key] = 0

                frame_index = frame_counters[key]
                frame_counters[key] += 1

                rows.append({
                    "run_folder": log_file.parent.name,
                    "target": target_name,
                    "line_number": line_number,
                    "channel": channel,
                    "exposure_s": exposures.get(channel),
                    "frame": frame_index,
                    "Cc": float(m.group("Cc")),
                    "Ca": float(m.group("Ca")),
                    "Nc": int(m.group("Nc")),
                    "Na": int(m.group("Na")),
                    "C_background": float(m.group("C_background")),
                    "C_star": float(m.group("C_star")),
                    "C_star_noise": float(m.group("C_star_noise")),
                })

    fieldnames = [
        "run_folder",
        "target",
        "line_number",
        "channel",
        "exposure_s",
        "frame",
        "Cc",
        "Ca",
        "Nc",
        "Na",
        "C_background",
        "C_star",
        "C_star_noise",
    ]

    with out_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: format_value_for_csv(v) for k, v in row.items()})

    csv_to_excel(out_csv, out_xlsx)

    print(f"Wrote {len(rows)} rows to {out_csv}")
    print(f"Wrote Excel file to {out_xlsx}")


if __name__ == "__main__":
    main("/Users/andreatanzer/Documents/Space Science/MasterThesis/WALTzER-simulator/output")