from astroquery.gaia import Gaia
from astropy.coordinates import SkyCoord
from astropy.table import Table
from datetime import datetime
from pathlib import Path
import sys
import numpy as np
import math
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table as XLTable
from openpyxl.worksheet.table import TableStyleInfo
from openpyxl.utils import get_column_letter

p = Path(__file__).resolve()
for parent in [p] + list(p.parents):
    if (parent / "src").exists() and (parent / "input").exists():
        src_path = parent / "src"
        if str(src_path) not in sys.path:
            sys.path.insert(0, str(src_path))
        break
else:
    raise RuntimeError("Repository root not found")

from loaders.run_waltzer_context import setup_output_directory
from loaders.run_waltzer_context import get_repo_root


def ts(*args):
    print(datetime.now().strftime("%H:%M:%S"), *args)


def is_missing(value):
    if value is None:
        return True
    if np.ma.is_masked(value):
        return True
    if isinstance(value, (float, np.floating)) and math.isnan(value):
        return True
    return False


def value_to_excel(value):
    if is_missing(value):
        return None
    return value


def format_source_id(source_id):
    s = str(source_id)
    groups = []
    while s:
        groups.append(s[-3:])
        s = s[:-3]
    return " ".join(reversed(groups))


def resolve_source_id(target_name, radius_arcsec=2.0):
    ts("Resolving target name...", target_name)

    coord = SkyCoord.from_name(target_name)
    ra0 = coord.ra.deg
    dec0 = coord.dec.deg
    radius_deg = radius_arcsec / 3600.0

    query = f"""
    SELECT TOP 1
        source_id,
        ra,
        dec,
        parallax,
        phot_g_mean_mag,
        DISTANCE(
            POINT('ICRS', ra, dec),
            POINT('ICRS', {ra0}, {dec0})
        ) * 3600.0 AS dist_arcsec
    FROM gaiadr3.gaia_source
    WHERE 1 = CONTAINS(
        POINT('ICRS', ra, dec),
        CIRCLE('ICRS', {ra0}, {dec0}, {radius_deg})
    )
    ORDER BY dist_arcsec ASC
    """

    job = Gaia.launch_job(query)
    result = job.get_results()

    if len(result) == 0:
        raise RuntimeError(f"no Gaia source found for {target_name}")

    source_id = int(result["source_id"][0])

    ts("source_id:", source_id)
    ts("distance_arcsec:", result["dist_arcsec"][0])

    return source_id


def get_source_id_and_label(star_name=None, star_id=None, radius_arcsec=2.0):
    if star_name is not None:
        source_id = resolve_source_id(star_name, radius_arcsec=radius_arcsec)
        label = star_name
        return source_id, label

    if star_id is not None:
        source_id = int(star_id)
        label = f"source_id_{source_id}"
        ts("Using provided source_id...", source_id)
        return source_id, label

    raise RuntimeError("Both star_name and star_id are None")


def query_table(source_id, table_name):
    ts("Querying", table_name)

    query = f"""
    SELECT *
    FROM {table_name}
    WHERE source_id = {source_id}
    """

    job = Gaia.launch_job(query)
    return job.get_results()


def collect_star_values(source_id, tables_to_check):
    star_values = {}

    for table_name in tables_to_check:
        try:
            tbl = query_table(source_id, table_name)

            if tbl is None or len(tbl) == 0:
                continue

            row = tbl[0]

            for col in tbl.colnames:
                key = (table_name, col)
                star_values[key] = value_to_excel(row[col])

        except Exception as e:
            ts("Failed for", table_name, ":", e)

    return star_values


def build_final_table(star_names, star_ids, tables_to_check):
    if star_names is None:
        star_names = []
    if star_ids is None:
        star_ids = []

    n_targets = max(len(star_names), len(star_ids))

    all_keys = set()
    star_column_names = []
    star_data = {}

    for i in range(n_targets):
        star_name = star_names[i] if i < len(star_names) else None
        star_id = star_ids[i] if i < len(star_ids) else None

        print()
        ts("Processing entry", i)

        try:
            source_id, label = get_source_id_and_label(star_name=star_name, star_id=star_id, radius_arcsec=2.0)
            source_id_fmt = format_source_id(source_id)
            star_col_name = f"{label} [{source_id_fmt}]"

            values = collect_star_values(source_id, tables_to_check)

            star_column_names.append(star_col_name)
            star_data[star_col_name] = values
            all_keys.update(values.keys())

        except Exception as e:
            ts("Failed for entry", i, ":", e)

    sorted_keys = sorted(all_keys, key=lambda x: (x[0], x[1]))

    rows = []
    for table_name, column_name in sorted_keys:
        row = {
            "gaia_table": table_name,
            "column_name": column_name,
        }

        for star_col in star_column_names:
            row[star_col] = star_data.get(star_col, {}).get((table_name, column_name), None)

        rows.append(row)

    colnames = ["gaia_table", "column_name"] + star_column_names
    final_table = Table(rows=rows, names=colnames)

    return final_table, star_column_names


def write_excel(final_table, output_file):
    df = final_table.to_pandas()

    df.to_excel(output_file, index=False, sheet_name="gaia_parameters")

    wb = load_workbook(output_file)
    ws = wb["gaia_parameters"]

    nrows = ws.max_row
    ncols = ws.max_column
    table_ref = f"A1:{get_column_letter(ncols)}{nrows}"

    excel_table = XLTable(displayName="GaiaParameters", ref=table_ref)
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    ws.add_table(excel_table)
    wb.save(output_file)


def main():
    Gaia.TIMEOUT = 120

    repo_root = get_repo_root(Path(__file__))
    print("repo_root =", repo_root)

    star_names = [
        "eps Boo",
        "WASP-12",
        "KELT-20",
        None,
        None,
        None,
        None,
    ]

    star_ids = [
        None,
        None,
        None,
        "2033110837908993920",
        "2033115442114430976",
        "2033124577474702336",
        "2033122795099056512",
    ]

    tables_to_check = [
        "gaiadr3.gaia_source",
        "gaiadr3.astrophysical_parameters",
        "gaiadr3.astrophysical_parameters_supp",
    ]

    final_table, _ = build_final_table(star_names, star_ids, tables_to_check)

    output_dir, _, _ = setup_output_directory()
    output_file = output_dir / "gaia_parameter_matrix.xlsx"

    write_excel(final_table, output_file)

    print()
    print("Saved Excel file to:", output_file)


if __name__ == "__main__":
    main()